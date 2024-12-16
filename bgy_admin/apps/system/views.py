from rest_framework import viewsets
from rest_framework.permissions import IsAuthenticated
from rest_framework.decorators import action
from drf_spectacular.utils import extend_schema, extend_schema_view, OpenApiParameter
from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment

from utils.decorators import log_operation
from utils.response import APIResponse

from .serializers import (
    RoleSerializer,
    RoleListSerializer,
    RoleDetailSerializer,
    RoleCreateSerializer,
    RoleUpdateSerializer
)
from apps.users.models import Role


@extend_schema_view(
    list=extend_schema(
        summary="获取角色列表",
        description="获取系统中的角色列表，支持分页和搜索",
        parameters=[
            OpenApiParameter(name='page', description='页码', required=False, type=int),
            OpenApiParameter(name='page_size', description='每页数量', required=False, type=int),
            OpenApiParameter(name='search', description='搜索关键词', required=False, type=str),
            OpenApiParameter(name='name', description='角色名称(模糊匹配)', required=False, type=str),
            OpenApiParameter(name='code', description='角色编码(模糊匹配)', required=False, type=str),
            OpenApiParameter(name='is_active', description='是否激活', required=False, type=bool),
            OpenApiParameter(name='ordering', description='排序字段(id/name/created_at)', required=False, type=str),
        ],
        responses={200: RoleListSerializer(many=True)}
    ),
    create=extend_schema(
        summary="创建角色",
        description="创建新角色，可以同时分配权限",
        request=RoleCreateSerializer,
        responses={201: RoleDetailSerializer}
    ),
    retrieve=extend_schema(
        summary="获取角色详情",
        description="获取指定角色的详细信息",
        parameters=[
            OpenApiParameter(name='id', description='角色ID', required=True, type=int, location=OpenApiParameter.PATH)
        ],
        responses={200: RoleDetailSerializer}
    ),
    update=extend_schema(
        summary="更新角色",
        description="更新指定角色的信息",
        parameters=[
            OpenApiParameter(name='id', description='角色ID', required=True, type=int, location=OpenApiParameter.PATH)
        ],
        request=RoleUpdateSerializer,
        responses={200: RoleDetailSerializer}
    ),
    destroy=extend_schema(
        summary="删除角色",
        description="删除指定角色(需确保没有用户使用该角色)",
        parameters=[
            OpenApiParameter(name='id', description='角色ID', required=True, type=int, location=OpenApiParameter.PATH)
        ],
        responses={204: None}
    )
)
@extend_schema(tags=['角色管理'])
class RoleViewSet(viewsets.ModelViewSet):
    """角色管理视图集"""
    queryset = Role.objects.all()
    serializer_class = RoleSerializer
    permission_classes = [IsAuthenticated]
    
    def get_serializer_class(self):
        """根据不同的action返回不同的序列化器"""
        if self.action == 'list':
            return RoleListSerializer
        elif self.action == 'create':
            return RoleCreateSerializer
        elif self.action in ['update', 'partial_update']:
            return RoleUpdateSerializer
        elif self.action in ['retrieve', 'assign_permissions']:
            return RoleDetailSerializer
        return self.serializer_class

    @log_operation('角色管理', '查询角色列表')
    def list(self, request, *args, **kwargs):
        return super().list(request, *args, **kwargs)

    @log_operation('角色管理', '创建角色')
    def create(self, request, *args, **kwargs):
        return super().create(request, *args, **kwargs)

    @log_operation('角色管理', '更新角色')
    def update(self, request, *args, **kwargs):
        return super().update(request, *args, **kwargs)

    @log_operation('角色管理', '删除角色')
    def destroy(self, request, *args, **kwargs):
        return super().destroy(request, *args, **kwargs)

    @extend_schema(
        summary="分配权限",
        description="为指定角色分配权限",
        parameters=[
            OpenApiParameter(name='id', description='角色ID', required=True, type=int, location=OpenApiParameter.PATH)
        ],
        request={
            'type': 'object',
            'properties': {
                'permissions': {
                    'type': 'array',
                    'items': {'type': 'integer'},
                    'description': '权限ID列表'
                }
            },
            'required': ['permissions']
        },
        responses={200: None}
    )
    @log_operation('角色管理', '分配权限')
    @action(detail=True, methods=['post'])
    def assign_permissions(self, request, pk=None):
        """分配权限"""
        role = self.get_object()
        permissions = request.data.get('permissions', [])
        role.permissions.set(permissions)
        return APIResponse(message="权限分配成功")

    @extend_schema(
        summary="导出角色",
        description="导出角色数据为Excel文件",
        responses={
            200: {
                'type': 'string',
                'format': 'binary',
                'content': {
                    'application/vnd.ms-excel': {}
                }
            }
        }
    )
    @log_operation('角色管理', '导出角色')
    @action(detail=False, methods=['get'])
    def export(self, request):
        """导出角色数据"""
        # 创建工作簿
        wb = Workbook()
        ws = wb.active
        ws.title = "角色列表"

        # 写入表头
        headers = ['角色名称', '角色编码', '状态', '创建时间', '备注']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center')

        # 写入数据
        roles = self.get_queryset()
        for row, role in enumerate(roles, 2):
            ws.cell(row=row, column=1, value=role.name)
            ws.cell(row=row, column=2, value=role.code)
            ws.cell(row=row, column=3, value='启用' if role.is_active else '禁用')
            ws.cell(row=row, column=4, value=role.created_at.strftime('%Y-%m-%d %H:%M:%S'))
            ws.cell(row=row, column=5, value=role.remark)

        # 调整列宽
        for col in ws.columns:
            max_length = 0
            for cell in col:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            ws.column_dimensions[col[0].column_letter].width = max_length + 2

        # 创建响应
        response = HttpResponse(
            content_type='application/vnd.ms-excel',
            headers={'Content-Disposition': 'attachment; filename="roles.xlsx"'},
        )
        wb.save(response)
        return response

    # ... 继续添加其他接口的文档 ...
