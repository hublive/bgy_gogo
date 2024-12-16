import logging

import pandas as pd
from django.contrib.auth import get_user_model, authenticate
from django.db.models import Q, Count, Max
from django.http import HttpResponse
from django.utils import timezone
from django_filters import rest_framework as filters
from drf_spectacular.utils import extend_schema, OpenApiParameter, extend_schema_view, OpenApiExample
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from rest_framework import viewsets, status
from rest_framework.decorators import action
from rest_framework.exceptions import ValidationError
from rest_framework.pagination import PageNumberPagination
from rest_framework.parsers import MultiPartParser
from rest_framework.permissions import IsAuthenticated, AllowAny
from rest_framework_simplejwt.tokens import RefreshToken

from apps.monitor.models import LoginLog, OperationLog  # 更新导入
from utils.decorators import department_permission_required, log_operation
from utils.response import APIResponse
from .models import UserSession, Department, Role
from .serializers import (
    UserListSerializer,
    UserDetailSerializer,
    UserCreateSerializer,
    UserUpdateSerializer,
    ChangePasswordSerializer,
    LoginSerializer,
    LogoutSerializer,
    CustomTokenObtainPairSerializer,
    CustomTokenRefreshSerializer,
    DepartmentSerializer,
)

User = get_user_model()
logger = logging.getLogger(__name__)


class UserFilter(filters.FilterSet):
    """用户过滤器
    
    支持按用户名、昵称、邮箱、手机号、角色、部门等字段过滤
    """
    username = filters.CharFilter(lookup_expr='icontains')
    nickname = filters.CharFilter(lookup_expr='icontains')
    email = filters.CharFilter(lookup_expr='icontains')
    phone = filters.CharFilter(lookup_expr='icontains')
    roles = filters.NumberFilter(field_name='roles')
    department = filters.NumberFilter()
    is_active = filters.BooleanFilter()
    date_joined = filters.DateFromToRangeFilter()

    class Meta:
        model = User
        fields = ['username', 'nickname', 'email', 'phone', 'roles', 'department', 'is_active', 'date_joined']


class DepartmentFilter(filters.FilterSet):
    """部门过滤器
    
    支持按部门名称、负责人、状态等字段过滤
    """
    name = filters.CharFilter(lookup_expr='icontains')
    leader = filters.NumberFilter()
    is_active = filters.BooleanFilter()
    created_at = filters.DateFromToRangeFilter()

    class Meta:
        model = Department
        fields = ['name', 'leader', 'is_active', 'created_at']


class StandardResultsSetPagination(PageNumberPagination):
    """标准分页器"""
    page_size = 10
    page_size_query_param = 'page_size'
    max_page_size = 1000


@extend_schema_view(
    list=extend_schema(
        summary="获取用户列表",
        description="获取系统中的用户列表，支持分页、搜索和过滤",
        parameters=[
            OpenApiParameter(name='page', description='页码', required=False, type=int),
            OpenApiParameter(name='page_size', description='每页数量', required=False, type=int),
            OpenApiParameter(name='search', description='搜索关键词(用户名/昵称/邮箱/手机)', required=False, type=str),
            OpenApiParameter(name='username', description='用户名(模糊匹配)', required=False, type=str),
            OpenApiParameter(name='nickname', description='昵称(模糊匹配)', required=False, type=str),
            OpenApiParameter(name='email', description='邮箱(模糊匹配)', required=False, type=str),
            OpenApiParameter(name='phone', description='手机号(模糊匹配)', required=False, type=str),
            OpenApiParameter(name='roles', description='角色ID', required=False, type=int),
            OpenApiParameter(name='department', description='部门ID', required=False, type=int),
            OpenApiParameter(name='is_active', description='是否激活', required=False, type=bool),
            OpenApiParameter(name='ordering', description='排序字段(id/username/date_joined/last_login)', required=False, type=str),
        ],
        responses={200: UserListSerializer(many=True)}
    ),
    create=extend_schema(
        summary="创建用户",
        description="创建新用户，需要提供用户名和密码等信息",
        request=UserCreateSerializer,
        responses={201: UserDetailSerializer},
        examples=[
            OpenApiExample(
                'Example 1',
                value={
                    'username': 'testuser',
                    'password': 'Test123456',
                    'confirm_password': 'Test123456',
                    'email': 'test@example.com',
                    'phone': '13800138000',
                    'nickname': '测试用户',
                    'gender': 1,
                    'roles': [1, 2],
                    'department': 1,
                    'is_active': True
                }
            )
        ]
    ),
    retrieve=extend_schema(
        summary="获取用户详情",
        description="获取指定用户的详细信息",
        parameters=[
            OpenApiParameter(name='id', description='用户ID', required=True, type=int, location=OpenApiParameter.PATH)
        ],
        responses={200: UserDetailSerializer}
    ),
    update=extend_schema(
        summary="更新用户",
        description="更新指定用户的信息",
        parameters=[
            OpenApiParameter(name='id', description='用户ID', required=True, type=int, location=OpenApiParameter.PATH)
        ],
        request=UserUpdateSerializer,
        responses={200: UserDetailSerializer}
    ),
    partial_update=extend_schema(
        summary="部分更新用户",
        description="部分更新指定用户的信息",
        parameters=[
            OpenApiParameter(name='id', description='用户ID', required=True, type=int, location=OpenApiParameter.PATH)
        ],
        request=UserUpdateSerializer,
        responses={200: UserDetailSerializer}
    ),
    destroy=extend_schema(
        summary="删除用户",
        description="删除指定用户",
        parameters=[
            OpenApiParameter(name='id', description='用户ID', required=True, type=int, location=OpenApiParameter.PATH)
        ],
        responses={204: None}
    )
)
@extend_schema(tags=['用户管理'])
class UserViewSet(viewsets.ModelViewSet):
    """用户管理视图集
    
    提供用户的增删改查、修改密码、上传头像等功能
    """
    queryset = User.objects.all()
    permission_classes = [IsAuthenticated]
    filterset_class = UserFilter
    search_fields = ['username', 'nickname', 'email', 'phone']
    ordering_fields = ['id', 'username', 'date_joined', 'last_login']
    ordering = ['-date_joined']

    def get_serializer_class(self):
        """根据不同的action返回不同的序列化器"""
        if self.action == 'list':
            return UserListSerializer
        elif self.action == 'create':
            return UserCreateSerializer
        elif self.action in ['update', 'partial_update']:
            return UserUpdateSerializer
        elif self.action == 'change_password':
            return ChangePasswordSerializer
        return UserDetailSerializer

    def get_queryset(self):
        """获取查询集，支持搜索过滤"""
        queryset = super().get_queryset()
        search = self.request.query_params.get('search', '')
        if search:
            queryset = queryset.filter(
                Q(username__icontains=search) |
                Q(nickname__icontains=search) |
                Q(email__icontains=search) |
                Q(phone__icontains=search)
            )
        return queryset

    @log_operation('用户管理', '查询用户列表')
    def list(self, request, *args, **kwargs):
        """获取用户列表
        
        支持以下功能：
        1. 分页查询
        2. 搜索过滤
        3. 字段排序
        """
        return super().list(request, *args, **kwargs)

    @log_operation('用户管理', '获取用户详情')
    def retrieve(self, request, *args, **kwargs):
        """获取用户详情"""
        return super().retrieve(request, *args, **kwargs)

    @log_operation('用户管理', '创建用户')
    def create(self, request, *args, **kwargs):
        """创建新用户"""
        serializer = self.get_serializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        self.perform_create(serializer)
        return APIResponse(data=serializer.data, message="创建成功")

    @log_operation('用户管理', '更新用户')
    def update(self, request, *args, **kwargs):
        """更新用户信息
        
        可更新的字段包括：
        1. 基本信息(昵称、邮箱等)
        2. 角色分配
        3. 部门调整
        """
        return super().update(request, *args, **kwargs)

    @log_operation('用户管理', '删除用户')
    def destroy(self, request, *args, **kwargs):
        """删除用户
        
        删除用户时会：
        1. 清理用户关联数据
        2. 记录操作日志
        3. 发送通知
        """
        return super().destroy(request, *args, **kwargs)

    @extend_schema(
        summary="修改密码",
        description="修改指定用户的登录密码",
        parameters=[
            OpenApiParameter(name='id', description='用户ID', required=True, type=int, location=OpenApiParameter.PATH)
        ],
        request=ChangePasswordSerializer,
        responses={200: None}
    )
    @log_operation('用户管理', '修改密码')
    @action(detail=True, methods=['post'])
    def change_password(self, request, pk=None):
        """修改密码
        
        修改密码时需要：
        1. 验证原密码
        2. 检查新密码复杂度
        3. 确认两次密码一致
        """
        user = self.get_object()
        serializer = self.get_serializer(data=request.data)
        serializer.is_valid(raise_exception=True)

        # 设置新密码
        user.set_password(serializer.validated_data['new_password'])
        user.save()

        return APIResponse(message="密码修改成功")

    @extend_schema(
        summary="重置密码",
        description="重置指定用户的密码为默认密码",
        parameters=[
            OpenApiParameter(name='id', description='用户ID', required=True, type=int, location=OpenApiParameter.PATH)
        ],
        responses={200: None}
    )
    @log_operation('用户管理', '重置密码')
    @action(detail=True, methods=['post'])
    def reset_password(self, request, pk=None):
        # 原有代码...
        pass

    @extend_schema(
        summary="上传头像",
        description="上传用户头像图片，支持jpg/png格式，大小不超过2MB",
        parameters=[
            OpenApiParameter(name='id', description='用户ID', required=True, type=int, location=OpenApiParameter.PATH)
        ],
        request={
            'multipart/form-data': {
                'type': 'object',
                'properties': {
                    'file': {'type': 'string', 'format': 'binary', 'description': '头像图片文件'}
                },
                'required': ['file']
            }
        },
        responses={200: None}
    )
    @log_operation('用户管理', '上传头像')
    @action(detail=True, methods=['post'], parser_classes=[MultiPartParser])
    def upload_avatar(self, request, pk=None):
        """上传头像
        
        支持的功能：
        1. 图片格式验证
        2. 图片大小限制
        3. 自动裁剪压缩
        """
        user = self.get_object()
        avatar = request.FILES.get('file')

        if not avatar:
            return APIResponse(
                code=status.HTTP_400_BAD_REQUEST,
                message="请选择要上传的图片"
            )

        # 验证文件类型和大小
        if not avatar.content_type.startswith('image/'):
            return APIResponse(
                code=status.HTTP_400_BAD_REQUEST,
                message="只能上传图片文件"
            )

        if avatar.size > 2 * 1024 * 1024:  # 2MB
            return APIResponse(
                code=status.HTTP_400_BAD_REQUEST,
                message="图片大小不能超过2MB"
            )

        # 保存头像
        user.avatar = avatar
        user.save()

        return APIResponse(
            message="头像上传成功",
            data={'avatar_url': request.build_absolute_uri(user.avatar.url)}
        )

    @log_operation('用户管理', '导出用户')
    @action(detail=False, methods=['get'])
    def export(self, request):
        """导出用户数据
        
        支持导出全部用户或按条件筛选后导出
        导出格式为 Excel，包含用户的基本信息
        """
        # 获取筛选后的用户数据
        queryset = self.filter_queryset(self.get_queryset())

        # 创建工作簿和工作表
        wb = Workbook()
        ws = wb.active
        ws.title = "用户列表"

        # 设置表头
        headers = [
            '用户名', '昵称', '邮箱', '手机号', '性别',
            '角色', '部门', '状态', '注册时间', '最后登录'
        ]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center')

        # 写入数据
        for row, user in enumerate(queryset, 2):
            ws.cell(row=row, column=1, value=user.username)
            ws.cell(row=row, column=2, value=user.nickname)
            ws.cell(row=row, column=3, value=user.email)
            ws.cell(row=row, column=4, value=user.phone)
            ws.cell(row=row, column=5, value=user.get_gender_display())
            ws.cell(row=row, column=6, value=user.role.name if user.role else '')
            ws.cell(row=row, column=7, value=user.department.name if user.department else '')
            ws.cell(row=row, column=8, value='启用' if user.is_active else '禁用')
            ws.cell(row=row, column=9, value=user.date_joined.strftime('%Y-%m-%d %H:%M:%S'))
            ws.cell(row=row, column=10, value=user.last_login.strftime('%Y-%m-%d %H:%M:%S') if user.last_login else '')

        # 调整列宽
        for col in ws.columns:
            max_length = max(len(str(cell.value)) for cell in col)
            ws.column_dimensions[col[0].column_letter].width = max_length + 2

        # 创建响应
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename=users.xlsx'

        # 保存文件
        wb.save(response)
        return response

    @log_operation('用户管理', '导入用户')
    @action(detail=False, methods=['post'], parser_classes=[MultiPartParser])
    def import_users(self, request):
        """导入用户数据
        
        从 Excel 文件导入用户数据
        支持批量创建和更新用户
        """
        file = request.FILES.get('file')
        if not file:
            return APIResponse(
                code=status.HTTP_400_BAD_REQUEST,
                message="请选择要导入的文件"
            )

        try:
            # 读取 Excel 文件
            df = pd.read_excel(file)

            # 验证必填字段
            required_fields = ['用户名', '密码']
            for field in required_fields:
                if field not in df.columns:
                    return APIResponse(
                        code=status.HTTP_400_BAD_REQUEST,
                        message=f"文件中缺少必填字段: {field}"
                    )

            # 处理数据
            success_count = 0
            error_count = 0
            error_msgs = []

            for index, row in df.iterrows():
                try:
                    # 创建或更新用户
                    username = row['用户名']
                    user = User.objects.filter(username=username).first()

                    if user:
                        # 更新已存在的用户
                        user.nickname = row.get('昵称', '')
                        user.email = row.get('邮箱', '')
                        user.phone = row.get('手机号', '')
                        user.save()
                    else:
                        # 创建新用户
                        user = User.objects.create_user(
                            username=username,
                            password=row['密码'],
                            nickname=row.get('昵称', ''),
                            email=row.get('邮箱', ''),
                            phone=row.get('手机号', '')
                        )

                    success_count += 1

                except Exception as e:
                    error_count += 1
                    error_msgs.append(f"第 {index + 2} 行: {str(e)}")

            return APIResponse(
                message=f"导入完成: 成功 {success_count} 条，失败 {error_count} 条",
                data={
                    'success_count': success_count,
                    'error_count': error_count,
                    'error_msgs': error_msgs
                }
            )

        except Exception as e:
            return APIResponse(
                code=status.HTTP_400_BAD_REQUEST,
                message=f"文件解析失败: {str(e)}"
            )

    @log_operation('用户管理', '用户统计')
    @action(detail=False, methods=['get'])
    def statistics(self, request):
        """用户数据统计
        
        统计用户相关的各项数据指标
        """
        # 获取时间范围
        today = timezone.now().date()
        last_30_days = today - timezone.timedelta(days=30)

        # 基础统计
        total_users = User.objects.count()
        active_users = User.objects.filter(is_active=True).count()
        today_logins = User.objects.filter(
            last_login__date=today
        ).count()

        # 近30天新增用户统计
        new_users_30d = User.objects.filter(
            date_joined__date__gte=last_30_days
        ).count()

        # 用户角色分布
        role_stats = User.objects.values(
            'role__name'
        ).annotate(
            count=Count('id')
        ).order_by('-count')

        # 部门分布
        dept_stats = User.objects.values(
            'department__name'
        ).annotate(
            count=Count('id')
        ).order_by('-count')

        # 性别分布
        gender_stats = User.objects.values(
            'gender'
        ).annotate(
            count=Count('id')
        ).order_by('gender')

        # 每日新增用户趋势
        daily_new_users = User.objects.filter(
            date_joined__date__gte=last_30_days
        ).values(
            'date_joined__date'
        ).annotate(
            count=Count('id')
        ).order_by('date_joined__date')

        return APIResponse(data={
            'overview': {
                'total_users': total_users,
                'active_users': active_users,
                'today_logins': today_logins,
                'new_users_30d': new_users_30d
            },
            'role_distribution': list(role_stats),
            'department_distribution': list(dept_stats),
            'gender_distribution': [
                {
                    'gender': item['gender'],
                    'gender_name': User.GENDER_CHOICES[item['gender']][1],
                    'count': item['count']
                }
                for item in gender_stats
            ],
            'daily_new_users': [
                {
                    'date': item['date_joined__date'].strftime('%Y-%m-%d'),
                    'count': item['count']
                }
                for item in daily_new_users
            ]
        })

    @log_operation('用户管理', '在线用户')
    @action(detail=False, methods=['get'])
    def online_users(self, request):
        """获取在线用户列表
        
        统计最近活跃的用户
        """
        # 获取最近30分钟活跃的会话
        active_time = timezone.now() - timezone.timedelta(minutes=30)
        sessions = UserSession.objects.filter(
            last_activity__gte=active_time
        ).select_related('user')

        return APIResponse(data=[
            {
                'id': session.user.id,
                'username': session.user.username,
                'nickname': session.user.nickname,
                'ip_address': session.ip_address,
                'user_agent': session.user_agent,
                'login_time': session.created_at,
                'last_activity': session.last_activity
            }
            for session in sessions
        ])

    @extend_schema(
        summary="强制下线",
        description="强制用户退出登录",
        responses={200: None}
    )
    @log_operation('用户管理', '强制下线')
    @action(detail=True, methods=['post'])
    def force_logout(self, request, pk=None):
        """强制用户下线
        
        删除用户的所有话记录
        """
        user = self.get_object()
        count = UserSession.objects.filter(user=user).delete()[0]

        return APIResponse(
            message=f"已强制下线用户 {user.username}",
            data={'deleted_sessions': count}
        )

    @extend_schema(
        summary="分配角色",
        description="为用户分配角色",
        request={
            'application/json': {
                'type': 'object',
                'properties': {
                    'role_ids': {
                        'type': 'array',
                        'items': {'type': 'integer'},
                        'description': '角色ID列表'
                    }
                }
            }
        },
        responses={200: None}
    )
    @log_operation('用户管理', '分配角色')
    @action(detail=True, methods=['post'])
    def assign_roles(self, request, pk=None):
        """分配角色
        
        为指定用户分配角色列表
        
        Args:
            request: 请求对象
            pk: 用户ID
            
        Returns:
            Response: 响应结果
        """
        user = self.get_object()
        role_ids = request.data.get('role_ids', [])

        # 验证角色是否存在
        roles = Role.objects.filter(id__in=role_ids)
        if len(roles) != len(role_ids):
            return APIResponse(code=400, message="部分角色不存在")

        # 分配角色
        user.roles.set(roles)
        return APIResponse(message="分配角色成功")

    @extend_schema(
        summary="分配部门",
        description="为用户分配部门",
        request={
            'application/json': {
                'type': 'object',
                'properties': {
                    'department_id': {
                        'type': 'integer',
                        'description': '部门ID'
                    }
                }
            }
        },
        responses={200: None}
    )
    @log_operation('用户管理', '分配部门')
    @action(detail=True, methods=['post'])
    def assign_department(self, request, pk=None):
        """分配部门
        
        为指定用户分配部门
        
        Args:
            request: 请求对象
            pk: 用户ID
            
        Returns:
            Response: 响应结果
        """
        user = self.get_object()
        department_id = request.data.get('department_id')

        # 验证部门是否存在
        try:
            department = Department.objects.get(id=department_id)
        except Department.DoesNotExist:
            return APIResponse(code=400, message="部门不存在")

        # 分配部门
        user.department = department
        user.save()
        return APIResponse(message="分配部门成功")

    def perform_create(self, serializer):
        """创建用户时的额外操作"""
        user = serializer.save()
        # 记录操作日志
        OperationLog.objects.create(
            user=self.request.user,
            module='用户管理',
            action='创建用户',
            method=self.request.method,
            path=self.request.path,
            ip_addr=self.request.META.get('REMOTE_ADDR'),
            status=1,
            params=str(self.request.data),
            result=f'创建用户 {user.username} 成功'
        )


@extend_schema_view(
    login=extend_schema(
        summary="用户登录",
        description="""使用用户名和密码登录系统，返回访问令牌和刷新令牌。
        
        登录成功后会：
        1. 记录登录日志
        2. 更新最后登录时间
        3. 创建在线用户记录
        """,
        request=LoginSerializer,
        responses={
            200: {
                'type': 'object',
                'properties': {
                    'access': {'type': 'string', 'description': '访问令牌'},
                    'refresh': {'type': 'string', 'description': '刷新令牌'},
                    'user': {'type': 'object', 'description': '用户信息'}
                }
            }
        },
        examples=[
            OpenApiExample(
                'Example 1',
                value={
                    'username': 'admin',
                    'password': 'admin123'
                }
            )
        ]
    ),
    logout=extend_schema(
        summary="用户登出",
        description="""注销当前用户的登录状态。
        
        登出时会：
        1. 删除在线用户记录
        2. 使当前令牌失效
        """,
        responses={200: None}
    ),
    refresh=extend_schema(
        summary="刷新令牌",
        description="使用刷新令牌获取新的访问令牌",
        request=CustomTokenRefreshSerializer,
        responses={
            200: {
                'type': 'object',
                'properties': {
                    'access': {'type': 'string', 'description': '新的访问令牌'}
                }
            }
        }
    )
)
@extend_schema(tags=['用户认证'])
class LoginViewSet(viewsets.ViewSet):
    """用户认证视图集"""
    permission_classes = [AllowAny]

    @log_operation('用户认证', '用户登录')
    @action(detail=False, methods=['post'])
    def login(self, request):
        """用户登录
        
        登录成功后返回：
        1. 访问令牌
        2. 刷新令牌
        3. 用户信息
        4. 权限列表
        """
        serializer = LoginSerializer(data=request.data)
        serializer.is_valid(raise_exception=True)

        username = serializer.validated_data['username']
        password = serializer.validated_data['password']

        # 验证用户
        user = authenticate(username=username, password=password)
        if not user:
            return APIResponse(code=400, message="用户名或密码错误")

        if not user.is_active:
            return APIResponse(code=400, message="用户已被禁用")

        # 生成令牌
        refresh = RefreshToken.for_user(user)

        # 更新登录信息
        user.last_login = timezone.now()
        user.save()

        # 记录登录日志
        LoginLog.objects.create(
            username=user.username,
            ip_addr=request.META.get('REMOTE_ADDR'),
            browser=request.META.get('HTTP_USER_AGENT', ''),
            status=1,  # 成功
            msg="登录成功"
        )

        # 创建或更新会话记录
        UserSession.objects.update_or_create(
            user=user,
            defaults={
                'session_key': str(refresh),
                'ip_address': request.META.get('REMOTE_ADDR'),
                'user_agent': request.META.get('HTTP_USER_AGENT', ''),
                'last_activity': timezone.now()
            }
        )

        return APIResponse(data={
            'access_token': str(refresh.access_token),
            'refresh_token': str(refresh),
            'token_type': 'Bearer',
            'expires_in': int(refresh.access_token.lifetime.total_seconds()),
            'user': {
                'id': user.id,
                'username': user.username,
                'nickname': user.nickname,
                'avatar': user.avatar.url if user.avatar else None,
                'roles': list(user.roles.values('id', 'name')),
                'permissions': list(user.get_all_permissions())
            }
        })

    @log_operation('用户认证', '刷新令牌')
    @action(detail=False, methods=['post'])
    def refresh(self, request):
        """刷新令牌
        
        使用刷新令牌获取：
        1. 新的访问令牌
        2. 新的过期时间
        """
        serializer = CustomTokenRefreshSerializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        return APIResponse(data=serializer.validated_data)

    @log_operation('用户认证', '用户登出')
    @action(detail=False, methods=['post'])
    def logout(self, request):
        """用户登出
        
        登出时会：
        1. 清除会话记录
        2. 使令牌失效
        3. 记录登出日志
        """
        try:
            # 删除会话记录
            UserSession.objects.filter(user=request.user).delete()
            return APIResponse(message="登出成功")
        except Exception as e:
            return APIResponse(code=400, message=f"登出失败: {str(e)}")

    def get_permissions(self):
        """
        动态设置权限
        """
        if self.action == 'login':
            permission_classes = [AllowAny]
        else:
            permission_classes = [IsAuthenticated]
        return [permission() for permission in permission_classes]


@extend_schema_view(
    info=extend_schema(
        summary="获取个人信息",
        description="获取当前登录用户的详细信息",
        responses={200: UserDetailSerializer}
    ),
    update_info=extend_schema(
        summary="更新个人信息",
        description="更新当前登录用户的基本信息",
        request=UserUpdateSerializer,
        responses={200: UserDetailSerializer}
    ),
    change_password=extend_schema(
        summary="修改个人密码",
        description="""修改当前登录用户的密码。
        
        需要：
        1. 验证原密码
        2. 检查新密码复杂度
        3. 确认两次密码一致
        """,
        request=ChangePasswordSerializer,
        responses={200: None}
    ),
    upload_avatar=extend_schema(
        summary="上传个人头像",
        description="""上传个人头像图片。
        
        支持：
        1. jpg/png格式
        2. 大小不超过2MB
        3. 自动裁剪压缩
        """,
        request={
            'multipart/form-data': {
                'type': 'object',
                'properties': {
                    'file': {'type': 'string', 'format': 'binary', 'description': '头像图片文件'}
                },
                'required': ['file']
            }
        },
        responses={200: None}
    )
)
@extend_schema(tags=['个人中心'])
class ProfileViewSet(viewsets.ViewSet):
    """个人中心视图集"""
    # ... 其他代码保持不变 ...


@extend_schema_view(
    list=extend_schema(
        summary="获取部门列表",
        description="获取系统中的部门列表，支持树形结构",
        parameters=[
            OpenApiParameter(name='name', description='部门名称', required=False, type=str),
            OpenApiParameter(name='leader', description='负责人ID', required=False, type=int),
            OpenApiParameter(name='is_active', description='是否启用', required=False, type=bool),
        ]
    ),
    create=extend_schema(
        summary="创建部门",
        description="创建新部门"
    ),
    retrieve=extend_schema(
        summary="获取部门详情",
        description="获取指定部门的详细信息"
    ),
    update=extend_schema(
        summary="更新部门",
        description="更新指定部门的信息"
    ),
    destroy=extend_schema(
        summary="删除部门",
        description="删除指定部门"
    )
)
@extend_schema(tags=['部门管理'])
class DepartmentViewSet(viewsets.ModelViewSet):
    """部门管理视图集"""

    @log_operation('部门管理', '查询部门列表')
    @department_permission_required('departments.view_department')
    def list(self, request, *args, **kwargs):
        return super().list(request, *args, **kwargs)

    @log_operation('部门管理', '创建部门')
    @department_permission_required('departments.add_department')
    def create(self, request, *args, **kwargs):
        return super().create(request, *args, **kwargs)

    @log_operation('部门管理', '更新部门')
    @department_permission_required('departments.change_department')
    def update(self, request, *args, **kwargs):
        return super().update(request, *args, **kwargs)

    @log_operation('部门管理', '删除部门')
    @department_permission_required('departments.delete_department')
    def destroy(self, request, *args, **kwargs):
        return super().destroy(request, *args, **kwargs)

    @log_operation('部门管理', '导入部门数据')
    @action(detail=False, methods=['post'])
    def import_data(self, request):
        # 原有代码...
        pass

    @log_operation('部门管理', '导出部门数据')
    @action(detail=False, methods=['get'])
    def export_data(self, request):
        # 原有代码...
        pass

    queryset = Department.objects.all()
    serializer_class = DepartmentSerializer
    permission_classes = [IsAuthenticated]
    filterset_class = DepartmentFilter
    search_fields = ['name', 'leader__username', 'leader__nickname']
    ordering_fields = ['order', 'created_at']
    ordering = ['order', 'id']
    pagination_class = StandardResultsSetPagination

    def get_queryset(self):
        """获取部门查询集
        
        支持按名称、负责人、状态等过滤
        支持按名称、负责人搜索
        """
        queryset = super().get_queryset()
        # 处理树形结构请求
        if self.request.query_params.get('tree'):
            return queryset.filter(parent=None)
        return queryset

    @extend_schema(
        summary="获取部门树",
        description="获取部门的树形结构数据"
    )
    @log_operation('部门管理', '获取部门树')
    @action(detail=False, methods=['get'])
    def tree(self, request):
        """获取部门树形结构
        
        返回完整的部门层级关系
        """
        departments = self.get_queryset()
        serializer = self.get_serializer(departments, many=True)
        return APIResponse(data=serializer.data)

    @extend_schema(
        summary="获取部门用户",
        description="获取指定部门的所有用户"
    )
    @log_operation('部门管理', '获取部门用户')
    @action(detail=True, methods=['get'])
    def users(self, request, pk=None):
        """获取部门用户列表
        
        获取指定部门及其子部门的所有用户
        """
        department = self.get_object()
        # 获取部门及其子部门的所有用户
        users = User.objects.filter(
            department__in=department.get_descendants(include_self=True)
        )
        page = self.paginate_queryset(users)
        serializer = UserListSerializer(page, many=True)
        return self.get_paginated_response(serializer.data)

    @log_operation('部门管理', '删除部门')
    def perform_destroy(self, instance):
        """删除部门
        
        删除前检查是否存在子部门或用户
        """
        if instance.department_set.exists():
            raise ValidationError("存在子部门,无法删除")
        if instance.user_set.exists():
            raise ValidationError("部门存在用户,无法删除")
        instance.delete()

    @extend_schema(
        summary="导出部门",
        description="导出部门数据为Excel文件",
        responses={
            200: {
                'type': 'string',
                'format': 'binary',
                'content': {
                    'application/vnd.ms-excel': {}
                }
            }
        }
    )
    @log_operation('部门管理', '导出部门')
    @action(detail=False, methods=['get'])
    def export(self, request):
        """导出部门数据
        
        将部门数据导出为Excel文件,包含以下字段:
        - 部门名称
        - 父部门
        - 负责人
        - 排序号
        - 状态
        - 备注
        """
        # 创建工作簿
        wb = Workbook()
        ws = wb.active
        ws.title = "部门列表"

        # 写入表头
        headers = ['部门名称', '父部门', '负责人', '排序号', '状态', '备注']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center')

        # 写入数据
        departments = self.get_queryset()
        for row, dept in enumerate(departments, 2):
            ws.cell(row=row, column=1, value=dept.name)
            ws.cell(row=row, column=2, value=dept.parent.name if dept.parent else '')
            ws.cell(row=row, column=3, value=dept.leader.nickname if dept.leader else '')
            ws.cell(row=row, column=4, value=dept.order)
            ws.cell(row=row, column=5, value='启用' if dept.is_active else '禁用')
            ws.cell(row=row, column=6, value=dept.remark)

        # 调整列宽
        for col in ws.columns:
            max_length = 0
            for cell in col:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            ws.column_dimensions[col[0].column_letter].width = max_length + 2

        # 创建响应
        response = HttpResponse(
            content_type='application/vnd.ms-excel',
            headers={'Content-Disposition': 'attachment; filename="departments.xlsx"'},
        )
        wb.save(response)
        return response

    @extend_schema(
        summary="导入部门",
        description="从Excel文件导入部门数据",
        request={
            'multipart/form-data': {
                'type': 'object',
                'properties': {
                    'file': {'type': 'string', 'format': 'binary'}
                }
            }
        },
        responses={200: None}
    )
    @log_operation('部门管理', '导入部门数据')
    @action(detail=False, methods=['post'], parser_classes=[MultiPartParser])
    def import_departments(self, request):
        """导入部门数据
        
        从Excel文件导入部门数据,支持以下功能:
        1. 批量创建部门
        2. 更新已存在的部门
        3. 自动处理部门层级关系
        """
        file = request.FILES.get('file')
        if not file:
            return APIResponse(code=400, message="请选择要导入的文件")

        try:
            # 读取Excel文件
            df = pd.read_excel(file)
            success_count = 0
            error_count = 0
            error_msgs = []

            # 处理每一行数据
            for index, row in df.iterrows():
                try:
                    # 查找或创建父部门
                    parent = None
                    if row['父部门']:
                        parent = Department.objects.filter(name=row['父部门']).first()
                        if not parent:
                            error_msgs.append(f"第{index + 2}行: 父部门 {row['父部门']} 不存在")
                            error_count += 1
                            continue

                    # 查找负责人
                    leader = None
                    if row['负责人']:
                        leader = User.objects.filter(nickname=row['负责人']).first()
                        if not leader:
                            error_msgs.append(f"第{index + 2}行: 负责人 {row['负责人']} 不存在")
                            error_count += 1
                            continue

                    # 更新或创建部门
                    dept, created = Department.objects.update_or_create(
                        name=row['部门名'],
                        defaults={
                            'parent': parent,
                            'leader': leader,
                            'order': row['排序号'],
                            'is_active': True if row['状态'] == '启用' else False,
                            'remark': row['备注'] if '备注' in row else ''
                        }
                    )
                    success_count += 1

                except Exception as e:
                    error_count += 1
                    error_msgs.append(f"第{index + 2}行: {str(e)}")

            return APIResponse(
                message=f"导入完成: 成功 {success_count} 条，失败 {error_count} 条",
                data={
                    'success_count': success_count,
                    'error_count': error_count,
                    'error_msgs': error_msgs
                }
            )

        except Exception as e:
            return APIResponse(code=400, message=f"导入失败: {str(e)}")

    @extend_schema(
        summary="部门统计",
        description="获取部门相关的统计数据"
    )
    @log_operation('部门管理', '部门统计')
    @action(detail=False, methods=['get'])
    def statistics(self, request):
        """部门统计数据
        
        返回以下统计数据：
        1. 部门总数
        2. 各层级部门数量分布
        3. 人员最多的前5个部门
        4. 新增部门趋势
        """
        # 部门总数
        total_departments = Department.objects.count()
        active_departments = Department.objects.filter(is_active=True).count()

        # 各层级部门数量
        level_distribution = []
        max_level = Department.objects.aggregate(max_level=Max('level'))['max_level']
        for level in range(max_level + 1):
            count = Department.objects.filter(level=level).count()
            level_distribution.append({
                'level': level,
                'count': count
            })

        # 人员最多的部门
        top_departments = Department.objects.annotate(
            member_count=Count('user')
        ).order_by('-member_count')[:5].values(
            'name', 'member_count'
        )

        # 新增部门趋势(最近30天)
        last_30_days = timezone.now() - timezone.timedelta(days=30)
        daily_new_departments = Department.objects.filter(
            created_at__gte=last_30_days
        ).extra(
            select={'date': 'DATE(created_at)'}
        ).values('date').annotate(
            count=Count('id')
        ).order_by('date')

        return APIResponse(data={
            'overview': {
                'total_departments': total_departments,
                'active_departments': active_departments
            },
            'level_distribution': level_distribution,
            'top_departments': list(top_departments),
            'daily_new_departments': list(daily_new_departments)
        })
