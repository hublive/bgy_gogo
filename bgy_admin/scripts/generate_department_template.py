from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

def generate_template():
    """生成部门导入模板
    
    创建一个包含示例数据的Excel模板文件
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "部门导入模板"
    
    # 设置列宽
    ws.column_dimensions['A'].width = 20  # 部门名称
    ws.column_dimensions['B'].width = 20  # 父部门
    ws.column_dimensions['C'].width = 15  # 负责人
    ws.column_dimensions['D'].width = 10  # 排序号
    ws.column_dimensions['E'].width = 10  # 状态
    ws.column_dimensions['F'].width = 30  # 备注
    
    # 设置样式
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(fill_type="solid", fgColor="366092")
    header_alignment = Alignment(horizontal="center", vertical="center")
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # 写入表头
    headers = ['部门名称*', '��部门', '负责人', '排序号', '状态', '备注']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = border
    
    # 写入示例数据
    example_data = [
        ['技术部', '', '张三', 1, '启用', '负责公司技术研发工作'],
        ['研发组', '技术部', '李四', 2, '启用', '负责具体研发工作'],
        ['测试组', '技术部', '王五', 3, '启用', '负责测试工作']
    ]
    
    for row, data in enumerate(example_data, 2):
        for col, value in enumerate(data, 1):
            cell = ws.cell(row=row, column=col, value=value)
            cell.alignment = Alignment(horizontal="center")
            cell.border = border
    
    # 添加说明
    ws = wb.create_sheet("填写说明")
    instructions = [
        ["部门导入模板填写说明"],
        [""],
        ["1. 带 * 的字段为必填项"],
        ["2. 父部门若为空,则表示一级部门"],
        ["3. 负责人请填写用户昵称"],
        ["4. 排序号为整数,值越小越靠前"],
        ["5. 状态只能填写: 启用/禁用"],
        ["6. 表格第一行(表头)请勿修改"],
        ["7. 示例数据仅供参考,可以删除"],
        [""],
        ["注意事项:"],
        ["1. 父部门必须已经存在"],
        ["2. 负责人必须是系统中已有的用户"],
        ["3. 同一层级下部门名称不能重复"],
        ["4. 部门层级不能超过5级"],
        ["5. 导入时会自动去除首尾空格"]
    ]
    
    for row, instruction in enumerate(instructions, 1):
        ws.cell(row=row, column=1, value=instruction[0])
    
    # 保存文件
    wb.save('部门导入模板.xlsx')

if __name__ == '__main__':
    generate_template() 